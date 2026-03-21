"""Smoke tests for DAZI Excel/spreadsheet processing.

Run with:
    cd /Users/wyatt/Downloads/DAZI/agent-operator
    python3 -m pytest apps/electron/resources/scripts/tests/test_xlsx_tool_smoke.py -v

Tests verify the xlsx skill under SKILLs/xlsx/ and core openpyxl operations
that the DAZI xlsx skill depends on.
"""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

try:
    from ._tool_test_harness import (
        SKILLS_DIR,
        XLSX_SCRIPTS,
        build_env,
        has_dependency,
        run_script,
    )
except ImportError:
    from _tool_test_harness import (
        SKILLS_DIR,
        XLSX_SCRIPTS,
        build_env,
        has_dependency,
        run_script,
    )


SKIP_REASON = None
if not has_dependency("openpyxl"):
    SKIP_REASON = "openpyxl not installed"


@unittest.skipIf(SKIP_REASON, SKIP_REASON or "")
class XlsxToolSmokeTests(unittest.TestCase):
    """Smoke tests for Excel/spreadsheet operations in DAZI."""

    @classmethod
    def setUpClass(cls) -> None:
        cls.env = build_env()
        cls.tmpdir_obj = tempfile.TemporaryDirectory(prefix="dazi-xlsx-smoke-")
        cls.tmpdir = Path(cls.tmpdir_obj.name)

        # Create a workbook with some data
        cls.book = cls.tmpdir / "workbook.xlsx"
        cls._create_test_workbook(cls.book)

    @classmethod
    def tearDownClass(cls) -> None:
        cls.tmpdir_obj.cleanup()

    @staticmethod
    def _create_test_workbook(path: Path) -> None:
        """Create a minimal test workbook using openpyxl."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "name"
        ws["B1"] = "score"
        ws["A2"] = "alice"
        ws["B2"] = 42
        ws["A3"] = "bob"
        ws["B3"] = 85
        wb.save(str(path))

    def test_write_read_roundtrip(self) -> None:
        """Verify basic write/read roundtrip with openpyxl (core DAZI xlsx dependency)."""
        from openpyxl import load_workbook

        wb = load_workbook(str(self.book))
        ws = wb.active

        self.assertEqual(ws["A1"].value, "name")
        self.assertEqual(ws["B1"].value, "score")
        self.assertEqual(ws["A2"].value, "alice")
        self.assertEqual(ws["B2"].value, 42)

    def test_add_sheet_and_info(self) -> None:
        """Verify adding a sheet and reading workbook metadata."""
        from openpyxl import load_workbook

        book_copy = self.tmpdir / "workbook_copy.xlsx"
        # Make a copy to avoid modifying shared fixture
        import shutil
        shutil.copy2(self.book, book_copy)

        wb = load_workbook(str(book_copy))
        self.assertGreaterEqual(len(wb.sheetnames), 1)

        # Add a new sheet
        wb.create_sheet("Data")
        wb.save(str(book_copy))

        wb2 = load_workbook(str(book_copy))
        self.assertIn("Data", wb2.sheetnames)

    def test_export_csv(self) -> None:
        """Verify exporting workbook data to CSV format."""
        import csv
        from openpyxl import load_workbook

        wb = load_workbook(str(self.book))
        ws = wb.active

        csv_path = self.tmpdir / "export.csv"
        with open(csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

        self.assertTrue(csv_path.exists())
        with open(csv_path) as f:
            reader = csv.reader(f)
            rows = list(reader)
        self.assertEqual(rows[0], ["name", "score"])
        self.assertEqual(rows[1][0], "alice")

    def test_invalid_sheet_errors(self) -> None:
        """Accessing a non-existent sheet should raise an error."""
        from openpyxl import load_workbook

        wb = load_workbook(str(self.book))
        with self.assertRaises(KeyError):
            _ = wb["NonExistentSheet"]

    def test_formula_support(self) -> None:
        """Verify that formulas can be written and read back."""
        from openpyxl import Workbook, load_workbook

        formula_book = self.tmpdir / "formulas.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=SUM(A1:A2)"
        wb.save(str(formula_book))

        wb2 = load_workbook(str(formula_book))
        ws2 = wb2.active
        self.assertEqual(ws2["A3"].value, "=SUM(A1:A2)")

    def test_recalc_script_exists(self) -> None:
        """Verify the recalc.py script is present in the xlsx skill."""
        recalc_script = XLSX_SCRIPTS / "recalc.py"
        self.assertTrue(
            recalc_script.exists(),
            f"Expected recalc script at {recalc_script}",
        )


if __name__ == "__main__":
    unittest.main()
